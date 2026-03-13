from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from langchain_community.document_loaders import PyPDFLoader
from langchain_community.document_loaders.word_document import Docx2txtLoader
from langchain_core.documents import Document

from app.enums import SupportedExtension


SUPPORTED_EXTENSIONS = {extension.value for extension in SupportedExtension}


def _base_metadata(
    *,
    doc_id: str,
    original_filename: str,
    stored_filename: str,
    stored_path: str,
    content_type: str | None,
    extension: str,
) -> dict:
    return {
        "doc_id": doc_id,
        "original_filename": original_filename,
        "stored_filename": stored_filename,
        "stored_path": stored_path,
        "content_type": content_type,
        "extension": extension,
    }


def _load_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _load_csv_as_text(path: Path) -> str:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file)
        return "\n".join(",".join("" if cell is None else str(cell) for cell in row) for row in reader)


def _load_xlsx_as_text(path: Path) -> Iterable[tuple[str, str]]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ModuleNotFoundError(
            "Missing dependency for .xlsx support. Install `openpyxl`."
        ) from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        lines: list[str] = []
        for row in worksheet.iter_rows(values_only=True):
            if not row:
                continue
            line = "\t".join("" if cell is None else str(cell) for cell in row).strip()
            if line:
                lines.append(line)
        yield sheet_name, "\n".join(lines)


def _load_xls_as_text(path: Path) -> Iterable[tuple[str, str]]:
    try:
        import xlrd  # type: ignore
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ModuleNotFoundError(
            "Missing dependency for .xls support. Install `xlrd`."
        ) from exc

    workbook = xlrd.open_workbook(path)
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        lines: list[str] = []
        for row_index in range(sheet.nrows):
            row = sheet.row_values(row_index)
            line = "\t".join("" if cell is None else str(cell) for cell in row).strip()
            if line:
                lines.append(line)
        yield sheet.name, "\n".join(lines)


def load_document(
    file_path: str,
    *,
    doc_id: str,
    original_filename: str,
    stored_filename: str,
    content_type: str | None = None,
) -> list[Document]:
    path = Path(file_path)
    extension = Path(original_filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file type: {extension}")

    base_metadata = _base_metadata(
        doc_id=doc_id,
        original_filename=original_filename,
        stored_filename=stored_filename,
        stored_path=str(path),
        content_type=content_type,
        extension=extension,
    )

    if extension == SupportedExtension.PDF.value:
        loader = PyPDFLoader(str(path))
        docs = loader.load()
        for doc in docs:
            doc.metadata = {**(doc.metadata or {}), **base_metadata}
        return docs

    if extension == SupportedExtension.DOCX.value:
        loader = Docx2txtLoader(str(path))
        docs = loader.load()
        for doc in docs:
            doc.metadata = {**(doc.metadata or {}), **base_metadata}
        return docs

    if extension in {
        SupportedExtension.TXT.value,
        SupportedExtension.MARKDOWN.value,
        SupportedExtension.MARKDOWN_ALT.value,
    }:
        return [
            Document(
                page_content=_load_text_file(path),
                metadata={**base_metadata, "source": str(path)},
            )
        ]

    if extension == SupportedExtension.CSV.value:
        return [
            Document(
                page_content=_load_csv_as_text(path),
                metadata={**base_metadata, "source": str(path)},
            )
        ]

    if extension == SupportedExtension.XLSX.value:
        documents: list[Document] = []
        for sheet_name, sheet_text in _load_xlsx_as_text(path):
            documents.append(
                Document(
                    page_content=sheet_text,
                    metadata={**base_metadata, "source": str(path), "sheet_name": sheet_name},
                )
            )
        return documents

    if extension == SupportedExtension.XLS.value:
        documents: list[Document] = []
        for sheet_name, sheet_text in _load_xls_as_text(path):
            documents.append(
                Document(
                    page_content=sheet_text,
                    metadata={**base_metadata, "source": str(path), "sheet_name": sheet_name},
                )
            )
        return documents

    raise ValueError(f"Unsupported file type: {extension}. Supported: {sorted(SUPPORTED_EXTENSIONS)}")
